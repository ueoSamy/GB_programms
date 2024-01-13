import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from tkinter import simpledialog


class ExcelEditor:
    COLUMN_NAMES = ["№", "Дата", "Сотрудник ОТК", "Коеффициент сложности", "Заказ клиента", 'Оборудование', 'Модель',
                    'Количество', 'Тип проверки', 'Время', 'Подразделения сборки', 'Результаты ОТК',
                    'Замечания (Комментарии)', 'Критические замечания']

    def __init__(self, root):
        # Инициализация переменных
        self.tree_frame = None
        self.labels = []  # <--- Добавлено определение
        self.input_entries = []  # <--- Добавлено определение
        self.comboboxes = []  # <--- Добавлено определение

        # Внутри метода __init__
        self.column_names = []  # Добавлен список для хранения наименований колонок
        self.tree = ttk.Treeview(self.tree_frame, show="headings", selectmode="extended")

        self.root = root
        self.root.title("Excel Editor")
        self.root.geometry("1024x800")
        self.root.resizable(False, False)

        # Удалить лист
        self.delete_sheet_button = tk.Button(root, text="Удалить лист", command=self.delete_sheet)
        self.delete_sheet_button.grid(row=len(self.labels) + len(self.input_entries), column=6, pady=3, sticky="w")

        # Добавлена кнопка для очистки полей
        self.clear_fields_button = tk.Button(root, text="Очистить поля", command=self.clear_fields)
        self.clear_fields_button.grid(row=len(self.labels) + len(self.input_entries), column=5, pady=3, sticky="w")

        # Добавлена кнопка для обновления файла
        self.refresh_file_button = tk.Button(root, text="Обновить файл", command=self.refresh_file)
        self.refresh_file_button.grid(row=len(self.labels) + len(self.input_entries), column=4, pady=3, sticky="w")

        # Иконка для окна
        self.root.iconbitmap('F:\geekgit\\rdw.ico')

        # Создание рабочего листа
        self.load_data()

        # Инициализация переменных
        self.current_row = 1  # Текущая строка для редактирования
        self.undo_stack = []  # Стек для отката изменений
        self.redo_stack = []  # Стек для обратного отката

        # Наименования строк
        self.labels = ["Строка 1", "Строка 2", "Строка 3", "Строка 4", "Строка 5",
                       "Строка 6", "Строка 7", "Строка 8", "Строка 9", "Строка 10", "Строка 11"]

        # Всплывающие списки и кнопки для каждой строки
        self.comboboxes = []

        for i, label_text in enumerate(self.labels):
            label = tk.Label(root, text=f"Наименование строки: {label_text}")
            label.grid(row=i, column=0, pady=3, sticky="w")

            combobox = ttk.Combobox(root, values=["Значение 1", "Значение 2", "Значение 3", "Значение 4", "Значение 5"])
            combobox.grid(row=i, column=1, pady=3, sticky="w")
            self.comboboxes.append(combobox)

        # Поля ввода
        self.input_entries = []
        for i in range(4):
            label = tk.Label(root, text=f"Наименование поля {i + 1}:")
            label.grid(row=len(self.labels) + i, column=0, pady=3, sticky="w")
            entry = tk.Entry(root)
            entry.grid(row=len(self.labels) + i, column=1, pady=3, sticky="w")
            self.input_entries.append(entry)

        # Кнопки управления данными
        self.add_data_button = tk.Button(root, text="Добавить данные", command=self.add_data_row)
        self.add_data_button.grid(row=len(self.labels) + len(self.input_entries), column=0, pady=3, sticky="w")

        self.delete_data_button = tk.Button(root, text="Удалить данные", command=self.delete_data_row)
        self.delete_data_button.grid(row=len(self.labels) + len(self.input_entries), column=1, pady=3, sticky="w")

        self.edit_data_button = tk.Button(root, text="Редактировать данные", command=self.edit_data_row)
        self.edit_data_button.grid(row=len(self.labels) + len(self.input_entries), column=2, pady=3, sticky="w")

        # Добавлена кнопка для перезаписи значений
        self.save_changes_button = tk.Button(root, text="Сохранить файл", command=self.save_changes)
        self.save_changes_button.grid(row=len(self.labels) + len(self.input_entries), column=3, pady=3, sticky="w")

        # Добавлена кнопка для отката изменений
        self.undo_button = tk.Button(root, text="Откат", command=self.undo_changes)
        self.undo_button.grid(row=len(self.labels) + len(self.input_entries), column=4, pady=3, sticky="w")

        # Добавлена кнопка для обратного отката
        self.redo_button = tk.Button(root, text="Обратный откат", command=self.redo_changes)
        self.redo_button.grid(row=len(self.labels) + len(self.input_entries), column=5, pady=3, sticky="w")

        # Добавлена кнопка для создания нового листа
        self.create_sheet_button = tk.Button(root, text="Добавить лист", command=self.create_new_sheet)
        self.create_sheet_button.grid(row=len(self.labels) + len(self.input_entries), column=6, pady=3, sticky="w")

        # Добавлено поле для отображения доступных листов
        self.sheet_var = tk.StringVar()
        self.sheet_var.set(self.workbook.sheetnames[0])
        self.sheet_menu = tk.OptionMenu(root, self.sheet_var, *self.workbook.sheetnames)
        self.sheet_menu.grid(row=len(self.labels) + len(self.input_entries), column=7, pady=3, sticky="w")

        # Таблица для отображения данных
        self.tree_frame = ttk.Frame(root)
        self.tree_frame.grid(row=len(self.labels) + len(self.input_entries) + 1, column=0, columnspan=8, rowspan=10,
                             pady=5, sticky="nsew")

        self.tree = ttk.Treeview(self.tree_frame, columns=tuple(f"Col{i}" for i in range(1, 15)), show="headings",
                                 selectmode="extended")
        for i in range(1, 15):
            self.tree.heading(f"Col{i}", text=f"Колонка {i}")
            self.tree.column(f"Col{i}", width=70, anchor="center")
        self.tree.grid(row=0, column=0, columnspan=8, rowspan=8, pady=5, sticky="nsew")

        self.tree_scroll = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        self.tree_scroll.grid(row=0, column=8, rowspan=8, pady=5, sticky="ns")
        self.tree.configure(yscrollcommand=self.tree_scroll.set)

        # Загрузка данных
        self.show_data()

    def refresh_file(self):
        try:
            self.workbook.save("data.xlsx")
            messagebox.showinfo("Уведомление", "Файл успешно обновлен.")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при обновлении файла: {e}")

    def clear_fields(self):
        for combobox in self.comboboxes:
            combobox.set("")
        for entry in self.input_entries:
            entry.delete(0, tk.END)

    def load_data(self):
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
            if file_path:
                self.workbook = load_workbook(file_path)
            else:
                self.workbook = Workbook()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при открытии файла: {e}")
            self.workbook = Workbook()

        self.sheet = self.workbook.active

    def show_data(self, event=None):
        self.clear_tree()
        selected_sheet = self.sheet_var.get()
        self.sheet = self.workbook[selected_sheet]

        # Получаем наименования колонок из заголовков листа
        self.column_names = [col.value for col in self.sheet[1]]

        # Удаляем все старые колонки Treeview
        for col in self.tree["columns"]:
            self.tree.column(col, anchor="w", width=0)
        self.tree["columns"] = tuple()

        # Создаем колонки Treeview
        for i, col_name in enumerate(self.column_names):
            col_id = f"Col{i + 1}"
            if i < len(self.COLUMN_NAMES):
                col_text = self.COLUMN_NAMES[i]
            else:
                col_text = f"Колонка {i + 1}"

            self.tree.heading(col_id, text=col_text, anchor="w")
            self.tree.column(col_id, width=70, anchor="w", stretch=False)
            self.tree["columns"] += (col_id,)

        # Отображаем данные
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=1):
            self.tree.insert("", tk.END, values=[row_num] + list(row))

    def delete_sheet(self):
        sheet_name = simpledialog.askstring("Удаление листа", "Введите название листа для удаления:")
        if sheet_name:
            try:
                self.workbook.remove(self.workbook[sheet_name])
                self.sheet_var.set(self.workbook.sheetnames[0])
                self.update_sheet_menu()
                self.show_data()
                messagebox.showinfo("Уведомление", f"Лист '{sheet_name}' успешно удален.")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при удалении листа: {e}")

    def update_sheet_menu(self):
        menu = self.sheet_menu["menu"]
        menu.delete(0, "end")
        for sheet_name in self.workbook.sheetnames:
            menu.add_command(label=sheet_name, command=lambda value=sheet_name: self.sheet_var.set(value))

    def clear_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def add_data(self, idx):
        selected_value = self.comboboxes[idx].get()
        if selected_value:
            row_data = [selected_value] + [entry.get() for entry in self.input_entries]
            self.undo_stack.append(list(self.sheet.iter_rows(values_only=True)))
            self.sheet.append(row_data)
            self.workbook.save("data.xlsx")
            self.show_data()

    def add_data_row(self):
        new_row = [combobox.get() for combobox in self.comboboxes] + [entry.get() for entry in self.input_entries]
        self.sheet.append(new_row)
        self.workbook.save("data.xlsx")
        self.show_data()

    def delete_data(self, row_indices):
        if row_indices:
            self.undo_stack.append(list(self.sheet.iter_rows(values_only=True)))
            for row_index in sorted(row_indices, reverse=True):
                self.sheet.delete_rows(row_index + 1)
            self.workbook.save("data.xlsx")
            self.show_data()

    def delete_data_row(self):
        selected_items = self.tree.selection()
        selected_rows = [self.tree.index(item) for item in selected_items]
        self.delete_data(selected_rows)

    def edit_data(self, idx):
        selected_item = self.tree.selection()
        if selected_item:
            row_index = self.tree.index(selected_item[0])
            self.current_row = row_index + 1
            self.comboboxes[idx].set(self.sheet.cell(row=self.current_row, column=1).value)
            for i, entry in enumerate(self.input_entries):
                entry.delete(0, tk.END)
                entry.insert(0, self.sheet.cell(row=self.current_row, column=i + 2).value)

    def edit_data_row(self):
        selected_item = self.tree.selection()
        if selected_item:
            self.edit_data(0)

    def save_changes(self):
        selected_item = self.tree.selection()
        if selected_item:
            self.current_row = self.tree.index(selected_item[0]) + 1
            row_data = [combobox.get() for combobox in self.comboboxes] + [entry.get() for entry in self.input_entries]
            for i, value in enumerate(row_data):
                self.sheet.cell(row=self.current_row, column=i + 1, value=value)
            self.workbook.save("data.xlsx")
            self.show_data()
            messagebox.showinfo("Уведомление", "Данные успешно сохранены.")

    def undo_changes(self):
        if self.undo_stack:
            last_changes = self.undo_stack.pop()
            self.redo_stack.append(
                list(self.sheet.iter_rows(values_only=True)))  # Добавлено в стек для обратного отката
            self.sheet.delete_rows(1, self.sheet.max_row)
            for row in last_changes:
                self.sheet.append(row)
            self.workbook.save("data.xlsx")
            self.show_data()
            messagebox.showinfo("Уведомление", "Изменения успешно отменены.")

    def redo_changes(self):
        if self.redo_stack:
            last_changes = self.redo_stack.pop()
            self.undo_stack.append(list(self.sheet.iter_rows(values_only=True)))  # Добавлено в стек для отката
            self.sheet.delete_rows(1, self.sheet.max_row)
            for row in last_changes:
                self.sheet.append(row)
            self.workbook.save("data.xlsx")
            self.show_data()
            messagebox.showinfo("Уведомление", "Изменения успешно восстановлены.")

    def create_new_sheet(self):
        new_sheet_name = self.ask_for_sheet_name()
        if new_sheet_name:
            self.workbook.create_sheet(title=new_sheet_name)
            self.sheet_var.set(new_sheet_name)
            self.show_data()
            messagebox.showinfo("Уведомление", f"Лист '{new_sheet_name}' успешно создан.")

    def ask_for_sheet_name(self):
        new_sheet_name = simpledialog.askstring("Введите название листа", "Введите название нового листа:")
        if new_sheet_name:
            return new_sheet_name
        else:
            messagebox.showwarning("Предупреждение", "Название листа не может быть пустым.")
            return None


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelEditor(root)
    root.mainloop()