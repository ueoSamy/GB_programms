import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QComboBox, QLineEdit, QPushButton, \
    QTableWidget, QTableWidgetItem, QVBoxLayout, QFileDialog, QHeaderView, QMenu, QInputDialog, QMessageBox
from openpyxl import Workbook, load_workbook

class ExcelEditor(QMainWindow):
    COLUMN_NAMES = ["№", "Дата", "Сотрудник ОТК", "Коеффициент сложности", "Заказ клиента", 'Оборудование', 'Модель',
                    'Количество', 'Тип проверки', 'Время', 'Подразделения сборки', 'Результаты ОТК',
                    'Замечания (Комментарии)', 'Критические замечания']

    def __init__(self):
        super().__init__()

        self.column_names = [] 
        self.current_row = 1  
        self.undo_stack = []  
        self.redo_stack = []  

        self.labels = ["Строка 1", "Строка 2", "Строка 3", "Строка 4", "Строка 5",
                       "Строка 6", "Строка 7", "Строка 8", "Строка 9", "Строка 10", "Строка 11"]

        self.comboboxes = []
        self.input_entries = []

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Excel Editor")
        self.setGeometry(100, 100, 1024, 800)

        self.central_widget = QVBoxLayout()

        self.load_data()

        self.delete_sheet_button = QPushButton("Удалить лист", self)
        self.delete_sheet_button.clicked.connect(self.delete_sheet)
        self.central_widget.addWidget(self.delete_sheet_button)

        self.clear_fields_button = QPushButton("Очистить поля", self)
        self.clear_fields_button.clicked.connect(self.clear_fields)
        self.central_widget.addWidget(self.clear_fields_button)

        self.refresh_file_button = QPushButton("Обновить файл", self)
        self.refresh_file_button.clicked.connect(self.refresh_file)
        self.central_widget.addWidget(self.refresh_file_button)

        self.create_sheet_button = QPushButton("Добавить лист", self)
        self.create_sheet_button.clicked.connect(self.create_new_sheet)
        self.central_widget.addWidget(self.create_sheet_button)

        self.sheet_var = "Sheet1"  # Первый лист по умолчанию

        self.tree = QTableWidget(self)
        self.tree.setColumnCount(15)
        self.tree.setHorizontalHeaderLabels(self.COLUMN_NAMES)
        self.tree.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.central_widget.addWidget(self.tree)

        self.show_data()

        central_widget = QVBoxLayout()
        central_widget.addLayout(self.central_widget)
        self.setCentralWidget(central_widget)

        self.show()

    def refresh_file(self):
        try:
            self.workbook.save("data.xlsx")
            QMessageBox.information(self, "Уведомление", "Файл успешно обновлен.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при обновлении файла: {e}")

    def clear_fields(self):
        for combobox in self.comboboxes:
            combobox.setCurrentIndex(-1)
        for entry in self.input_entries:
            entry.clear()

    def load_data(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "Открыть файл", "", "Excel files (*.xlsx *.xls)")
            if file_path:
                self.workbook = load_workbook(file_path)
            else:
                self.workbook = Workbook()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при открытии файла: {e}")
            self.workbook = Workbook()

        self.sheet = self.workbook.active

    def show_data(self):
        self.tree.clearContents()
        self.tree.setRowCount(0)

        selected_sheet = self.sheet_var
        self.sheet = self.workbook[selected_sheet]

        self.column_names = [col.value for col in self.sheet[1]]

        self.tree.setHorizontalHeaderLabels(self.column_names)

        for row_num, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=1):
            self.tree.insertRow(row_num - 1)
            for col_num, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.tree.setItem(row_num - 1, col_num, item)

    def delete_sheet(self):
        sheet_name, ok = QInputDialog.getText(self, "Удаление листа", "Введите название листа для удаления:")
        if ok and sheet_name:
            try:
                self.workbook.remove(self.workbook[sheet_name])
                self.sheet_var = self.workbook.sheetnames[0]
                self.update_sheet_menu()
                self.show_data()
                QMessageBox.information(self, "Уведомление", f"Лист '{sheet_name}' успешно удален.")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении листа: {e}")

    def update_sheet_menu(self):
        # Дополнительные изменения не требуются, так как QComboBox автоматически обновляется

    def clear_tree(self):
        self.tree.clearContents()
        self.tree.setRowCount(0)

    def add_data(self, idx):
        selected_value = self.comboboxes[idx].currentText()
        if selected_value:
            row_data = [selected_value] + [entry.text() for entry in self.input_entries]
            self.undo_stack.append(list(self.sheet.iter_rows(values_only=True)))
            self.sheet.append(row_data)
            self.workbook.save("data.xlsx")
            self.show_data()

    def add_data_row(self):
        new_row = [combobox.currentText() for combobox in self.comboboxes] + [entry.text() for entry in self.input_entries]
        self.sheet.append(new_row)
        self.workbook.save("data.xlsx")
        self.show_data()

    def delete_data(self, row_indices):
        if row_indices:
            self.undo_stack.append(list(self.sheet.iter_rows(values_only=True)))
            for row_index in sorted(row_indices, reverse=True):
                self.sheet.delete_rows(row_index + 1)
            self.workbook.save("data.xlsx")
            self.show_data()

    def delete_data_row(self):
        selected_rows = [item.row() for item in self.tree.selectedItems()]
        self.delete_data(selected_rows)

    def edit_data(self, idx):
        selected_item = self.tree.selectedItems()
        if selected_item:
            row_index = selected_item[0].row()
            self.current_row = row_index + 1
            self.comboboxes[idx].setCurrentText(str(self.sheet.cell(row=self.current_row, column=1).value))
            for i, entry in enumerate(self.input_entries):
                entry.clear()
                entry.insert(0, str(self.sheet.cell(row=self.current_row, column=i + 2).value))

    def edit_data_row(self):
        selected_item = self.tree.selectedItems()
        if selected_item:
            self.edit_data(0)

    def save_changes(self):
        selected_item = self.tree.selectedItems()
        if selected_item:
            self.current_row = selected_item[0].row() + 1
            row_data = [combobox.currentText() for combobox in self.comboboxes] + [entry.text() for entry in self.input_entries]
            for i, value in enumerate(row_data):
                self.sheet.cell(row=self.current_row, column=i + 1, value=value)
            self.workbook.save("data.xlsx")
            self.show_data()
            QMessageBox.information(self, "Уведомление", "Данные успешно сохранены.")

    def undo_changes(self):
        if self.undo_stack:
            last_changes = self.undo_stack.pop()
            self.redo_stack.append(list(self.sheet.iter_rows(values_only=True)))
            self.sheet.delete_rows(1, self.sheet.max_row)
            for row in last_changes:
                self.sheet.append(row)
            self.workbook.save("data.xlsx")
            self.show_data()
            QMessageBox.information(self, "Уведомление", "Изменения успешно отменены.")

    def redo_changes(self):
        if self.redo_stack:
            last_changes = self.redo_stack.pop()
            self.undo_stack.append(list(self.sheet.iter_rows(values_only=True)))
            self.sheet.delete_rows(1, self.sheet.max_row)
            for row in last_changes:
                self.sheet.append(row)
            self.workbook.save("data.xlsx")
            self.show_data()
            QMessageBox.information(self, "Уведомление", "Изменения успешно восстановлены.")

    def create_new_sheet(self):
        new_sheet_name, ok = QInputDialog.getText(self, "Введите название листа", "Введите название нового листа:")
        if ok and new_sheet_name:
            self.workbook.create_sheet(title=new_sheet_name)
            self.sheet_var = new_sheet_name
            self.show_data()
            QMessageBox.information(self, "Уведомление", f"Лист '{new_sheet_name}' успешно создан.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = ExcelEditor()
    sys.exit(app.exec_())
